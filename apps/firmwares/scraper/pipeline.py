"""
pipeline.py v5 — Streaming multi-format output pipeline.

New in v5:
  • ParquetWriter  — Apache Parquet via pyarrow (optional dep)
  • DataEnricher   — structured numeric fields on every record
  • ValidationLayer— per-record validation with warning/error counts
  • BrandProgressTracker — per-brand scraping progress
  • DeadLetterQueue — exhausted-retry URLs written to dead_letter.json
  • DedupFilter v2  — bloom filter + model-name dedup
  • CLI-controlled output formats (from Settings.effective_output_formats)
  • Per-format flush intervals tunable separately

Strict-mode Pylance: Settings typed directly (not Any), TextIOWrapper explicit.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import sqlite3
import time
from pathlib import Path
from typing import Any

from .dedup import BrandProgressTracker, DeadLetterQueue, DedupFilter
from .enricher import DataEnricher
from .profile_engine import ProfileEngine
from .settings import Settings

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column ordering
# ---------------------------------------------------------------------------

_PRIORITY: list[str] = [
    # Identity
    "brand",
    "brand_slug",
    "model_name",
    "full_name",
    "url",
    "image_url",
    "model_variants_str",
    "model_variant_count",
    # Derived numeric (enricher)
    "launch_year",
    "launch_quarter",
    "display_size_in",
    "weight_g_num",
    "battery_mah_num",
    "ram_gb_num",
    "storage_gb_num",
    "main_camera_mp_num",
    "price_usd_approx",
    "network_gen_derived",
    # v6 enriched boolean flags
    "has_5g",
    "has_nfc",
    "has_esim",
    "has_headphone_jack",
    "has_dual_speakers",
    "has_ir_blaster",
    "has_wifi6",
    # v6 enriched structured fields
    "chipset_family",
    "launch_status_clean",
    "form_factor_clean",
    "ip_rating",
    "ip_water_depth_m",
    "review_score",
    "charging_speed_w",
    "refresh_rate_hz",
    "screen_to_body_pct",
    "cpu_cores",
    "bluetooth_version",
    "wifi_generation",
    "popularity_rank",
    # Platform
    "platform_os",
    "os_version",
    "platform_chipset",
    "chipset",
    "platform_cpu",
    "cpu",
    "platform_gpu",
    "gpu",
    "chipset_process_node",
    "chipset_cpu_arch",
    "chipset_fab",
    "chipset_chipset_year",
    # Launch
    "launch_announced",
    "launch_status",
    # Network
    "network_technology",
    "network_gen",
    "network_2g_bands",
    "network_3g_bands",
    "network_4g_bands",
    "network_5g_bands",
    # Body
    "body_dimensions",
    "dimensions",
    "body_weight",
    "weight_g",
    "body_build",
    "body_sim",
    # Display
    "display_type",
    "screen_type",
    "display_size",
    "screen_size_inches",
    "display_resolution",
    "screen_resolution",
    "display_protection",
    # Memory
    "memory_card_slot",
    "memory_internal",
    "storage_ram",
    # Cameras
    "main_camera_single",
    "main_camera_dual",
    "main_camera_triple",
    "main_camera_quad",
    "main_camera_features",
    "main_camera_video",
    "selfie_camera_single",
    "selfie_camera_dual",
    "selfie_camera_features",
    "selfie_camera_video",
    # Sound
    "sound_loudspeaker",
    "sound_jack_35mm",
    # Comms
    "comms_wlan",
    "comms_bluetooth",
    "comms_positioning",
    "comms_nfc",
    "comms_radio",
    "comms_usb",
    # Features
    "features_sensors",
    # Battery
    "battery_type",
    "battery_capacity",
    "battery_mah",
    "battery_charging",
    # Misc
    "misc_colors",
    "misc_models",
    "variant_codes_raw",
    "misc_sar",
    "misc_sar_eu",
    "misc_price",
    "price_eur",
    # GSMArena meta
    "gsmarena_rating",
    "gsmarena_fans",
    "review_url",
    "video_review_url",
    "pros_cons",
    "review_snippet",
    "related_devices",
    "is_tablet",
    "_source",
    "scraped_at",
]


def _ordered_keys(record: dict[str, Any], extra: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    all_keys = set(record.keys()) | set(extra)
    for col in _PRIORITY:
        if col in all_keys and col not in seen:
            out.append(col)
            seen.add(col)
    for col in sorted(all_keys):
        if col not in seen:
            out.append(col)
            seen.add(col)
    return out


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v)


def _safe_col(name: str) -> str:
    return re.sub(r"[^a-z0-9_]", "_", name.lower())


# ---------------------------------------------------------------------------
# CsvWriter
# ---------------------------------------------------------------------------


class CsvWriter:
    def __init__(self, path: str, resume: bool) -> None:
        self._path = path
        self._mode = "a" if resume else "w"
        self._file: io.TextIOWrapper | None = None
        self._writer: csv.DictWriter[str] | None = None
        self._initial_fields: list[str] = []
        self._all_keys: set[str] = set()
        self._count: int = 0

    def _open(self, record: dict[str, Any]) -> None:
        self._initial_fields = list(_ordered_keys(record, []))
        self._all_keys = set(self._initial_fields)
        f = open(self._path, self._mode, newline="", encoding="utf-8")  # noqa: SIM115
        assert isinstance(f, io.TextIOWrapper)
        self._file = f
        self._writer = csv.DictWriter(
            self._file,
            fieldnames=list(self._initial_fields),
            extrasaction="ignore",
            lineterminator="\n",
        )
        p = Path(self._path)
        if self._mode == "w" or not p.exists() or p.stat().st_size == 0:
            self._writer.writeheader()

    def write(self, record: dict[str, Any]) -> None:
        if self._file is None:
            self._open(record)
        self._all_keys.update(record.keys())
        if self._writer is None:
            raise RuntimeError("CsvWriter._writer is None after _open()")
        self._writer.writerow({k: _to_str(v) for k, v in record.items()})
        self._count += 1
        if self._count % 50 == 0 and self._file is not None:
            self._file.flush()

    def consolidate(self, records: list[dict[str, Any]]) -> None:
        if not records:
            return
        if self._file is not None:
            try:
                self._file.close()
            except OSError as exc:
                logger.warning("CSV incremental close error: %s", exc)
            finally:
                self._file = None
                self._writer = None

        all_keys: set[str] = set()
        for r in records:
            all_keys.update(r.keys())
        fieldnames = _ordered_keys(records[0], list(all_keys))
        with open(self._path, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(
                fh,
                fieldnames=fieldnames,
                extrasaction="ignore",
                lineterminator="\n",
            )
            w.writeheader()
            for r in records:
                w.writerow({k: _to_str(v) for k, v in r.items()})
        logger.info(
            "CSV consolidated: %d records × %d cols → %s",
            len(records),
            len(fieldnames),
            self._path,
        )

    def close(self) -> None:
        if self._file is not None:
            try:
                self._file.flush()
                self._file.close()
            except OSError as exc:
                logger.warning("CSV close error: %s", exc)
            finally:
                self._file = None
        logger.info("CsvWriter closed: %d records → %s", self._count, self._path)


# ---------------------------------------------------------------------------
# JsonlWriter
# ---------------------------------------------------------------------------


class JsonlWriter:
    def __init__(self, path: str, resume: bool) -> None:
        self._path = path
        self._file: io.TextIOWrapper = open(
            path, "a" if resume else "w", encoding="utf-8"
        )
        self._count: int = 0

    def write(self, record: dict[str, Any]) -> None:
        self._file.write(json.dumps(record, ensure_ascii=False, default=_to_str) + "\n")
        self._count += 1
        if self._count % 50 == 0:
            self._file.flush()

    def close(self) -> None:
        try:
            self._file.flush()
            self._file.close()
        except OSError as exc:
            logger.warning("JSONL close error: %s", exc)
        logger.info("JsonlWriter closed: %d records → %s", self._count, self._path)


# ---------------------------------------------------------------------------
# SqliteWriter
# ---------------------------------------------------------------------------


class SqliteWriter:
    def __init__(self, path: str) -> None:
        self._path = path
        self._conn: sqlite3.Connection | None = None
        self._known: set[str] = {"id", "url"}
        self._count: int = 0

    def _connect(self) -> None:
        self._conn = sqlite3.connect(self._path, check_same_thread=False)
        self._conn.execute("PRAGMA journal_mode=WAL;")
        self._conn.execute("PRAGMA synchronous=NORMAL;")
        self._conn.execute(
            "CREATE TABLE IF NOT EXISTS devices "
            "(id INTEGER PRIMARY KEY AUTOINCREMENT, url TEXT UNIQUE NOT NULL);"
        )
        self._conn.commit()
        for row in self._conn.execute("PRAGMA table_info(devices);").fetchall():
            self._known.add(row[1])

    def _ensure_cols(self, record: dict[str, Any]) -> None:
        if self._conn is None:
            raise RuntimeError("SqliteWriter not connected")
        for key in record:
            sc = _safe_col(key)
            if sc not in self._known:
                try:
                    self._conn.execute(f"ALTER TABLE devices ADD COLUMN [{sc}] TEXT;")
                    self._known.add(sc)
                except sqlite3.OperationalError:
                    self._known.add(sc)

    def write(self, record: dict[str, Any]) -> None:
        if self._conn is None:
            self._connect()
        self._ensure_cols(record)
        if self._conn is None:
            raise RuntimeError("SqliteWriter._conn None after _connect()")
        safe = {_safe_col(k): _to_str(v) for k, v in record.items()}
        cols = ", ".join(f"[{c}]" for c in safe)
        placeholders = ", ".join("?" for _ in safe)
        try:
            self._conn.execute(
                f"INSERT OR REPLACE INTO devices ({cols}) VALUES ({placeholders});",  # noqa: S608
                list(safe.values()),
            )
            self._count += 1
            if self._count % 100 == 0:
                self._conn.commit()
        except sqlite3.Error as exc:
            logger.warning("SQLite write error: %s", exc)

    def close(self) -> None:
        if self._conn is None:
            return
        try:
            self._conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_brand ON devices (brand);"
            )
            self._conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_launch ON devices (launch_year);"
            )
            self._conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_net ON devices (network_gen_derived);"
            )
        except sqlite3.Error as exc:
            logger.debug("SQLite index creation: %s", exc)
        try:
            self._conn.commit()
            self._conn.close()
        except sqlite3.Error as exc:
            logger.warning("SQLite close error: %s", exc)
        logger.info("SqliteWriter closed: %d records → %s", self._count, self._path)


# ---------------------------------------------------------------------------
# ExcelWriter
# ---------------------------------------------------------------------------


class ExcelWriter:
    def __init__(self, path: str) -> None:
        self._path = path
        self._rows: list[dict[str, Any]] = []
        try:
            import openpyxl  # type: ignore[import-untyped]

            _ = openpyxl
            self._ok = True
        except ImportError:
            logger.info("openpyxl not installed — Excel output disabled")
            self._ok = False

    def write(self, record: dict[str, Any]) -> None:
        if self._ok:
            self._rows.append(record)

    def close(self) -> None:
        if not self._ok or not self._rows:
            return
        try:
            import openpyxl  # type: ignore[import-untyped]
            from openpyxl.styles import (  # type: ignore[import-untyped]
                Font,
                PatternFill,
            )

            all_keys: set[str] = set()
            for r in self._rows:
                all_keys.update(r.keys())
            headers = _ordered_keys(self._rows[0], list(all_keys))

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:  # pragma: no cover — Workbook always has active sheet
                ws = wb.create_sheet("Devices")
            ws.title = "Devices"

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                )
            ws.freeze_panes = "A2"

            for ri, record in enumerate(self._rows, 2):
                for ci, key in enumerate(headers, 1):
                    ws.cell(row=ri, column=ci, value=_to_str(record.get(key, "")))

            for col_cells in ws.columns:
                max_len = max(
                    (len(_to_str(c.value)) for c in col_cells if c.value is not None),
                    default=10,
                )
                first_cell = col_cells[0]
                col_letter = getattr(first_cell, "column_letter", "A")
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

            wb.save(self._path)
            logger.info("Excel saved: %d records → %s", len(self._rows), self._path)
        except Exception as exc:
            logger.error("Excel write failed: %s", exc)


# ---------------------------------------------------------------------------
# ParquetWriter (NEW in v5)
# ---------------------------------------------------------------------------


class ParquetWriter:
    """
    Apache Parquet output via pyarrow.
    Columnar format — ideal for pandas / Spark / DuckDB analytics.
    Written at finalize() (not streaming) because Parquet requires full schema upfront.
    Falls back gracefully if pyarrow not installed.
    """

    def __init__(self, path: str) -> None:
        self._path = path
        self._rows: list[dict[str, Any]] = []
        try:
            import pyarrow  # type: ignore[import-untyped]

            _ = pyarrow
            self._ok = True
        except ImportError:
            logger.info(
                "pyarrow not installed — Parquet output disabled. "
                "Run: pip install pyarrow"
            )
            self._ok = False

    def write(self, record: dict[str, Any]) -> None:
        if self._ok:
            self._rows.append(record)

    def close(self) -> None:
        if not self._ok or not self._rows:
            return
        try:
            import pyarrow as pa  # type: ignore[import-untyped]
            import pyarrow.parquet as pq  # type: ignore[import-untyped]

            all_keys: set[str] = set()
            for r in self._rows:
                all_keys.update(r.keys())
            headers = _ordered_keys(self._rows[0], list(all_keys))

            # Build column arrays
            columns: dict[str, list[object]] = {h: [] for h in headers}
            for r in self._rows:
                for h in headers:
                    v = r.get(h)
                    columns[h].append(
                        _to_str(v)
                        if not isinstance(v, (int, float, bool, type(None)))
                        else v
                    )

            # Infer types: numeric enricher fields should be int/float, rest strings
            _FLOAT_COLS = {
                "display_size_in",
                "ram_gb_num",
                "storage_gb_num",
                "main_camera_mp_num",
                "price_usd_approx",
            }
            _INT_COLS = {
                "weight_g_num",
                "battery_mah_num",
                "launch_year",
                "model_variant_count",
            }

            arrays: list[Any] = []
            fields: list[Any] = []
            for h in headers:
                col = columns[h]
                if h in _INT_COLS:
                    safe_col = [
                        int(str(v)) if v not in (None, "", "None") else None
                        for v in col
                    ]
                    arr = pa.array(safe_col, type=pa.int32())
                    fields.append(pa.field(h, pa.int32()))
                elif h in _FLOAT_COLS:
                    safe_col = [
                        float(str(v)) if v not in (None, "", "None") else None
                        for v in col
                    ]
                    arr = pa.array(safe_col, type=pa.float32())
                    fields.append(pa.field(h, pa.float32()))
                else:
                    arr = pa.array(
                        [str(v) if v is not None else "" for v in col], type=pa.string()
                    )
                    fields.append(pa.field(h, pa.string()))
                arrays.append(arr)

            schema = pa.schema(fields)
            table = pa.table(dict(zip(headers, arrays, strict=False)), schema=schema)
            pq.write_table(table, self._path, compression="snappy")
            logger.info(
                "Parquet saved: %d records × %d cols → %s",
                len(self._rows),
                len(headers),
                self._path,
            )
        except Exception as exc:
            logger.error("Parquet write failed: %s", exc)


# ---------------------------------------------------------------------------
# CrawlStats
# ---------------------------------------------------------------------------


class CrawlStats:
    def __init__(self) -> None:
        self._t0: float = time.monotonic()
        self.total_scraped: int = 0
        self.total_failed: int = 0
        self.total_banned: int = 0
        self.total_validation_warnings: int = 0
        self.total_validation_errors: int = 0
        self.per_brand: dict[str, int] = {}

    def record(self, brand: str) -> None:
        self.total_scraped += 1
        self.per_brand[brand] = self.per_brand.get(brand, 0) + 1

    @property
    def elapsed(self) -> float:
        return time.monotonic() - self._t0

    @property
    def rate(self) -> float:
        e = self.elapsed
        return self.total_scraped / e if e > 0 else 0.0

    def eta(self, total: int | None) -> str:
        if not total or self.rate == 0:
            return "?"
        secs = max(0.0, (total - self.total_scraped) / self.rate)
        return f"{secs / 3600:.1f}h" if secs > 3600 else f"{secs / 60:.1f}m"

    def summary(self, total: int | None = None) -> list[str]:
        lines = [
            f"  Scraped  : {self.total_scraped}",
            f"  Failed   : {self.total_failed}",
            f"  Banned   : {self.total_banned}",
            f"  Val warns: {self.total_validation_warnings}",
            f"  Val errs : {self.total_validation_errors}",
            f"  Rate     : {self.rate:.2f}/s",
            f"  Elapsed  : {self.elapsed / 60:.1f}min",
        ]
        if total:
            pct = self.total_scraped / total * 100
            lines += [
                f"  Progress : {pct:.1f}% ({self.total_scraped}/{total})",
                f"  ETA      : {self.eta(total)}",
            ]
        top = sorted(self.per_brand.items(), key=lambda x: -x[1])[:5]
        if top:
            lines.append("  Top brands: " + ", ".join(f"{b}:{c}" for b, c in top))
        return lines


# ---------------------------------------------------------------------------
# OutputPipeline
# ---------------------------------------------------------------------------


class OutputPipeline:
    """
    Orchestrates all writers + profile engine + enricher + validation.

    Pipeline per device:
      dedup → profile filter → enrich → validate → write all formats
    """

    def __init__(self, cfg: Settings, profile: ProfileEngine, job: Any = None) -> None:
        self._cfg = cfg
        self._profile = profile
        self._enricher = DataEnricher()
        self._job = job
        self._start_time = __import__("time").monotonic()

        # Resolve active output formats (CLI can override profile)
        self._formats = cfg.effective_output_formats(profile.output_formats)

        self.dedup = DedupFilter(
            bloom=cfg.bloom_filter,
            capacity=cfg.bloom_capacity,
            error_rate=cfg.bloom_error_rate,
            dedup_by_model=cfg.dedup_by_model,
        )
        self.stats = CrawlStats()
        self.brand_progress = BrandProgressTracker()
        self.dead_letter = DeadLetterQueue(cfg.output_dir)

        self._records: list[dict[str, Any]] = []
        self._failures: list[dict[str, str]] = []
        self._validation_log: list[dict[str, Any]] = []

        self._csv: CsvWriter | None = (
            CsvWriter(cfg.output_csv, cfg.resume) if self._formats.get("csv") else None
        )
        self._jsonl: JsonlWriter | None = (
            JsonlWriter(cfg.output_jsonl, cfg.resume)
            if self._formats.get("jsonl")
            else None
        )
        self._sqlite: SqliteWriter | None = (
            SqliteWriter(cfg.output_sqlite) if self._formats.get("sqlite") else None
        )
        self._excel: ExcelWriter | None = (
            ExcelWriter(cfg.output_excel) if self._formats.get("excel") else None
        )
        self._parquet: ParquetWriter | None = (
            ParquetWriter(cfg.output_parquet) if self._formats.get("parquet") else None
        )

        if cfg.resume:
            self.dedup.seed_from_csv(cfg.output_csv)

    # -- Public API ----------------------------------------------------------

    def process(self, raw_item: dict[str, Any]) -> bool:
        """Dedup → profile → enrich → validate → write. Returns True if written."""
        url = str(raw_item.get("url", ""))
        brand = str(raw_item.get("brand", ""))
        model = str(raw_item.get("model_name", ""))

        if not url or self.dedup.is_seen(url, brand, model):
            return False
        self.dedup.mark(url, brand, model)

        # Profile filter
        item = self._profile.process(raw_item)

        # Enrichment — adds numeric / date / price fields
        item = self._enricher.enrich(item)

        # Validation
        if self._cfg.validate_on_write:
            vr = self._enricher.validate(item)
            self.stats.total_validation_warnings += len(vr.warnings)
            self.stats.total_validation_errors += len(vr.errors)
            if vr.warnings or vr.errors:
                self._validation_log.append(
                    {
                        "url": url,
                        "warnings": vr.warnings,
                        "errors": vr.errors,
                    }
                )
            if not vr.is_valid:
                logger.warning(
                    "Validation failed [%s %s]: %s",
                    brand,
                    model,
                    "; ".join(vr.errors),
                )

        self._write(item)
        self._records.append(item)
        self.stats.record(brand)
        self.brand_progress.record_scraped(brand)

        n = self.stats.total_scraped
        if n % self._cfg.flush_every == 0:
            logger.info(
                "Pipeline: %d written | %.2f/s | %d brands | warnings=%d",
                n,
                self.stats.rate,
                len(self.stats.per_brand),
                self.stats.total_validation_warnings,
            )
        return True

    async def process_async(self, raw_item: dict[str, Any]) -> bool:
        """Async variant — runs AI enrichment via profile_engine."""
        url = str(raw_item.get("url", ""))
        brand = str(raw_item.get("brand", ""))
        model = str(raw_item.get("model_name", ""))

        if not url or self.dedup.is_seen(url, brand, model):
            return False
        self.dedup.mark(url, brand, model)

        item = await self._profile.process_async(raw_item)
        item = self._enricher.enrich(item)

        if self._cfg.validate_on_write:
            vr = self._enricher.validate(item)
            self.stats.total_validation_warnings += len(vr.warnings)
            self.stats.total_validation_errors += len(vr.errors)

        self._write(item)
        self._records.append(item)
        self.stats.record(brand)
        self.brand_progress.record_scraped(brand)
        return True

    def record_failure(self, url: str, reason: str, brand: str = "") -> None:
        self.stats.total_failed += 1
        if brand:
            self.brand_progress.record_failed(brand)
        self._failures.append({"url": url, "reason": reason})

    def record_dead_letter(self, url: str, reason: str, attempt: int) -> None:
        self.dead_letter.record(url, reason, attempt)

    def finalize(self) -> None:
        if self._csv:
            self._csv.consolidate(self._records)
            self._csv.close()
        if self._jsonl:
            self._jsonl.close()
        if self._sqlite:
            self._sqlite.close()
        if self._excel:
            self._excel.close()
        if self._parquet:
            self._parquet.close()
        self._write_json()
        self._write_failures()
        self._write_validation_log()
        self._write_brand_progress()
        self._write_job_manifest()
        self.dead_letter.finalize()
        self._profile.finalize()
        logger.info(
            "Pipeline finalized: %d devices | %d warnings | %d errors",
            len(self._records),
            self.stats.total_validation_warnings,
            self.stats.total_validation_errors,
        )

    # -- Internal ------------------------------------------------------------

    def _write(self, item: dict[str, Any]) -> None:
        if self._csv:
            self._csv.write(item)
        if self._jsonl:
            self._jsonl.write(item)
        if self._sqlite:
            self._sqlite.write(item)
        if self._excel:
            self._excel.write(item)
        if self._parquet:
            self._parquet.write(item)

    def _write_json(self) -> None:
        if not self._formats.get("json"):
            return
        try:
            with open(self._cfg.output_json, "w", encoding="utf-8") as fh:
                json.dump(
                    self._records,
                    fh,
                    ensure_ascii=False,
                    indent=2,
                    default=_to_str,
                )
            logger.info(
                "JSON written: %d records → %s",
                len(self._records),
                self._cfg.output_json,
            )
        except OSError as exc:
            logger.error("JSON write failed: %s", exc)

    def _write_failures(self) -> None:
        if not self._failures:
            return
        p = Path(self._cfg.output_dir) / "failures.json"
        try:
            with open(p, "w", encoding="utf-8") as fh:
                json.dump(self._failures, fh, indent=2, ensure_ascii=False)
            logger.info("Failures: %d → %s", len(self._failures), p)
        except OSError as exc:
            logger.error("Failures write failed: %s", exc)

    def _write_validation_log(self) -> None:
        if not self._validation_log:
            return
        p = Path(self._cfg.output_dir) / "validation_log.json"
        try:
            with open(p, "w", encoding="utf-8") as fh:
                json.dump(self._validation_log, fh, indent=2, ensure_ascii=False)
            logger.info(
                "Validation log: %d records with issues → %s",
                len(self._validation_log),
                p,
            )
        except OSError as exc:
            logger.error("Validation log write failed: %s", exc)

    def _write_brand_progress(self) -> None:
        p = Path(self._cfg.output_dir) / "brand_progress.json"
        try:
            with open(p, "w", encoding="utf-8") as fh:
                json.dump(self.brand_progress.report(), fh, indent=2)
            logger.info("Brand progress → %s", p)
        except OSError as exc:
            logger.error("Brand progress write failed: %s", exc)

    def _write_job_manifest(self) -> None:
        """Write a machine-readable manifest of this crawl run."""
        import time as _time

        elapsed = _time.monotonic() - self._start_time
        manifest: dict[str, Any] = {
            "scraper_version": "6.0",
            "strategy": self._cfg.crawl_strategy,
            "job_id": (getattr(self._job, "job_id", None) if self._job else None),
            "total_scraped": self.stats.total_scraped,
            "total_failed": self.stats.total_failed,
            "total_banned": self.stats.total_banned,
            "dead_letter": len(self.dead_letter),
            "duration_seconds": round(elapsed, 1),
            "output_dir": self._cfg.output_dir,
            "active_formats": {k: v for k, v in self._formats.items() if v},
        }
        if self._job:
            try:
                manifest["job"] = self._job.to_dict()
            except Exception:  # noqa: S110
                pass
        p = Path(self._cfg.output_dir) / "job_manifest.json"
        try:
            with open(p, "w", encoding="utf-8") as fh:
                json.dump(manifest, fh, indent=2, ensure_ascii=False)
            logger.info("Job manifest → %s", p)
        except OSError as exc:
            logger.error("Job manifest write failed: %s", exc)
